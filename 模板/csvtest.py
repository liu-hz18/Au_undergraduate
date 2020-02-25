
import csv
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import math
import numpy as np
import matplotlib.pyplot as plt

def csv_test():
    headers =  ['class', 'name', 'sex', 'height', 'year']
    rows_list = [
        [1, 'xiaoming', 'male', 168, 23],
        [1, 'xiaoming', 'female', 162, 22],
        [2, 'xiaozhang', 'female', 163, 21],
        [2, 'xiaoli', 'male', 158, 21]
    ]
    rows_dict = [
        {'class': 1, 'name': 'xiaoming', 'sex': 'male', 'height': 168, 'year': 3},
        {'class': 1, 'name': 'xiaohong', 'sex': 'female', 'height': 162, 'year': 22},
        {'class': 2, 'name': 'xiaozhang', 'sex': 'female', 'height': 163, 'year': 21},
        {'class': 2, 'name': 'xiaoli', 'sex': 'male', 'height': 158, 'year': 1},
    ]
    # 文件写入，字典和列表形式
    with open('./test1.csv', 'w', newline='') as f:
        csvfile = csv.writer(f)
        csvfile.writerow(headers)# 写入表头
        csvfile.writerows(rows_list)# 写入内容
    with open('test2.csv', 'w', newline='')as f:
        f_csv = csv.DictWriter(f, headers)
        f_csv.writeheader()
        f_csv.writerows(rows_dict)
    # 文件读取
    with open('test1.csv', 'r') as f:
        f_csv = csv.reader(f)
        print(f_csv)
        for row in f_csv:
            print(row)# row是一个列表
    with open('test1.csv', 'r') as f:
        f_csv = csv.DictReader(f)# 以字典形式读入
        print(f_csv)
        for row in f_csv:
            print(row) # row是一个OrderedDict
            print(row['class'])


def excel_test():
    # 取出所有工作簿
    wb = openpyxl.load_workbook('./dataProcess.xlsx')# 不支持xls的读入
    # 获得表单
    sheets = wb.sheetnames
    print(sheets, type(sheets))
    # 创建新表单
    mySheet = wb.create_sheet('mySheet')
    print(mySheet)
    print(wb.sheetnames)
    # 获取表单
    sheet2 = wb['Sheet1']
    print(sheet2)
    # 打印表单名字
    for sheet in wb:
        print(sheet.title)
    # 获取单元格信息
    ws = wb.active # 当前活跃的表单，即上一次关闭时光标所在表单页
    print(ws)
    print(ws['A1'])# 获取A列的第一个对象
    print(ws['A1'].value)
    print(ws.cell(row=1, column=1))
    print(ws.cell(row=1, column=1).value)

    c = ws['B1']#Cell 对象
    print('Row {}, Column {} is {}'.format(c.row, c.column, c.value)) # 打印这个单元格对象所在的行列的数值和内容
    print('Cell {} is {}\n'.format(c.coordinate, c.value)) # 获取单元格对象的所在列的行数和值

    row = ws[1] # 第1行，序从1开始
    column = ws['A'] # 第A列,里面存的是 Cell 对象
    print(column)

    row_range = ws[2:6]
    col_range = ws['B:C']

    # 把数字转换成字母
    print(get_column_letter(2), get_column_letter(47), get_column_letter(900))
    # 把字母转换成数字
    print(column_index_from_string('AAH'))
    wb.save(filename='./data.xlsx')#建议保存成xlsx, 不保存则之前的所有操作都不会被硬盘记录


class Data:
    def __init__(self, name, code, GNIppp, life, gini="0.0"):
        self.name = name
        self.code = code
        self.GNIppp = int(GNIppp)
        self.life = float(life)
        self.gini = float(gini)

    def _print_(self):
        print("Country:" + self.name + "  Code:" + self.code + "  GNI:" + str(self.GNIppp) + "  Life:" + str(self.life) + "  Gini:" + str(self.gini))



def worldDataClean():
    # 清洗世界银行数据，发展经济学课程作业
    workbook = openpyxl.load_workbook('./GNI_LIFE_2017.xlsx')
    #workbook = openpyxl.load_workbook('./GINI_LIFE_GNI_2010.xlsx')
    dataSheet = workbook['Data']
    CountryName = dataSheet['A'][1:]
    CountryCode = dataSheet['B'][1:]
    '''
    GNIPPP = dataSheet['K'][1:]
    expectedLife = dataSheet['H'][1:]
    gini = dataSheet['E'][1:]
    size = len(CountryName)
    dataArray = []
    print(size)
    for i in range(size):
        #print(gini[i].value)
        if GNIPPP[i].value != '..' and expectedLife[i].value != '..' and gini[i].value != '..':
            dataArray.append(Data(CountryName[i].value, CountryCode[i].value, GNIPPP[i].value, expectedLife[i].value, gini[i].value))
    size = len(dataArray)
    '''
    
    GNIPPP = dataSheet['E'][1:]
    expectedLife = dataSheet['F'][1:]
    size = len(CountryName)
    dataArray = []
    print(size)
    for i in range(size):
        #print(gini[i].value)
        if GNIPPP[i].value != '..' and expectedLife[i].value != '..':
            dataArray.append(Data(CountryName[i].value, CountryCode[i].value, GNIPPP[i].value, expectedLife[i].value))
    size = len(dataArray)
    
    GNIArray = []
    lifeArray = []
    CountryName = []
    gini = []

    china = 0
    america = 0
    farest = 0
    i = 0
    for data in dataArray:
        data._print_()
        if data.name == 'China':
            china = i
        if data.name == "United States":
            america = i
        if data.name == "Equatorial Guinea":
            farest = i
        GNIArray.append(data.GNIppp)
        lifeArray.append(data.life)
        CountryName.append(data.name)
        gini.append(data.gini)
        i += 1
    conuntryNum = i #有效数据个数
    
    # printPlot(np.array(GNIArray), np.array(lifeArray), CountryName, china, america, conuntryNum)
    printPlot(np.log(GNIArray), np.array(lifeArray), CountryName, china, america, conuntryNum)
    
    '''
    printPlot(np.array(GNIArray), np.array(gini), CountryName, china, america, conuntryNum)
    printPlot(np.log(GNIArray), np.array(gini), CountryName, china, america, conuntryNum)
    printPlot(np.array(gini), np.array(lifeArray), CountryName, china, america, conuntryNum)
    '''

#线性拟合
def printPlot(x, y, name, china, america, num):
    f1 = np.polyfit(x, y, 1)
    print(f1)
    new_x = np.arange(min(x), max(x), (max(x) - min(x))/100000)# 设置取点间隔
    y_curve = np.polyval(f1, new_x)# 拟合y值
    plot1 = plt.plot(x, y, '.', label='original values', markersize=5)
    # plot3 = plt.plot(x, yvals, '.', label='predicted values')
    plot2 = plt.plot(new_x, y_curve, ',', label='log regression', color='r')
    # plt.text(1.7, 1.8, "r="+str(r)[:6])
    # plt.text(1.7, 2.0, "y="+str(a)[:5]+"x+"+str(b)[:5])
    #for i in range(len(name)):
    #    plt.text(x[i], y[i], name[i])
    plt.text(x[china], y[china], 'China')
    plt.text(x[america], y[america], 'United States')
    plt.xlabel('ln(GNI per capita(PPP$))')
    plt.ylabel('Life expectancy at birth')
    plt.legend(loc=4)#指定legend的位置右下角
    plt.title('\'GNI per capita(PPP$)(ln) - Life expectancy at birth\'(2017)(Valid Country:%d)' % num)#名称还没改
    plt.scatter(x[china], y[china], color='r')
    plt.scatter(x[america], y[america], color='r')
    plt.show()



if __name__ == '__main__':
    # csv_test()
    # excel_test()
    worldDataClean()
