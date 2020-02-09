
import csv
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def csv_test():
    headers =  ['class', 'name', 'sex', 'height', 'year']
    rows_list = [
        [1, 'xiaoming', 'male', 168, 23],
        [1, 'xiaoming', 'female', 162, 22],
        [2, 'xiaozhang', 'female', 163, 21],
        [2, 'xiaoli', 'male', 158, 21]
    ]
    rows_dict = [
        {'class': 1, 'name': 'xiaoming', 'sex': 'male', 'height': 168, 'year': 3},
        {'class': 1, 'name': 'xiaohong', 'sex': 'female', 'height': 162, 'year': 22},
        {'class': 2, 'name': 'xiaozhang', 'sex': 'female', 'height': 163, 'year': 21},
        {'class': 2, 'name': 'xiaoli', 'sex': 'male', 'height': 158, 'year': 1},
    ]
    # 文件写入，字典和列表形式
    with open('./test1.csv', 'w', newline='') as f:
        csvfile = csv.writer(f)
        csvfile.writerow(headers)# 写入表头
        csvfile.writerows(rows_list)# 写入内容
    with open('test2.csv', 'w', newline='')as f:
        f_csv = csv.DictWriter(f, headers)
        f_csv.writeheader()
        f_csv.writerows(rows_dict)
    # 文件读取
    with open('test1.csv', 'r') as f:
        f_csv = csv.reader(f)
        print(f_csv)
        for row in f_csv:
            print(row)# row是一个列表
    with open('test1.csv', 'r') as f:
        f_csv = csv.DictReader(f)# 以字典形式读入
        print(f_csv)
        for row in f_csv:
            print(row) # row是一个OrderedDict
            print(row['class'])


def excel_test():
    # 取出所有工作簿
    wb = openpyxl.load_workbook('./dataProcess.xlsx')# 不支持xls的读入
    # 获得表单
    sheets = wb.sheetnames
    print(sheets, type(sheets))
    # 创建新表单
    mySheet = wb.create_sheet('mySheet')
    print(mySheet)
    print(wb.sheetnames)
    # 获取表单
    sheet2 = wb['Sheet1']
    print(sheet2)
    # 打印表单名字
    for sheet in wb:
        print(sheet.title)
    # 获取单元格信息
    ws = wb.active # 当前活跃的表单，即上一次关闭时光标所在表单页
    print(ws)
    print(ws['A1'])# 获取A列的第一个对象
    print(ws['A1'].value)
    print(ws.cell(row=1, column=1))
    print(ws.cell(row=1, column=1).value)

    c = ws['B1']#Cell 对象
    print('Row {}, Column {} is {}'.format(c.row, c.column, c.value)) # 打印这个单元格对象所在的行列的数值和内容
    print('Cell {} is {}\n'.format(c.coordinate, c.value)) # 获取单元格对象的所在列的行数和值

    row = ws[1] # 第1行，序从1开始
    column = ws['A'] # 第A列,里面存的是 Cell 对象
    print(column)

    row_range = ws[2:6]
    col_range = ws['B:C']

    # 把数字转换成字母
    print(get_column_letter(2), get_column_letter(47), get_column_letter(900))
    # 把字母转换成数字
    print(column_index_from_string('AAH'))
    wb.save(filename='./data.xlsx')#建议保存成xlsx, 不保存则之前的所有操作都不会被硬盘记录


if __name__ == '__main__':
    csv_test()
    excel_test()
